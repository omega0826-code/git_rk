#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
설문조사 주관식 의견 분석 스크립트
경북TP 공동활용 연구장비 관련 설문조사 주관식 의견을 카테고리별로 분류하고 Excel 보고서 생성
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re

# 카테고리 정의 및 키워드 매핑
CATEGORIES = {
    '홍보 및 정보 제공': [
        '홍보', '정보', '인지', '알', '모르', '안내', '소개', '접근성', 
        '플랫폼', '사이트', '홈페이지', 'SNS', '신문', '책자'
    ],
    '장비 관리 및 운영': [
        '고장', '노후', '수리', '유지보수', '관리', '담당자', '인원', 
        '부족', '업무', '가동', '운영', '안정성'
    ],
    '사용 절차 및 편의성': [
        '절차', '간소', '신청', '예약', '번거', '편의', '접근', 
        '이용', '사용', '활용', '대응'
    ],
    '장비 및 서비스 확충': [
        '장비', '증설', '도입', '신규', '추가', '확보', '다양', 
        '보유', '리스트', '종류', '성능'
    ],
    '전문 지원 및 교육': [
        '분석', '해석', '전문', '교육', '매뉴얼', '지원', '설명', 
        '안내', '결과', '이해'
    ],
    '비용 및 지원 사업': [
        '비용', '지원', '사용료', '금액', '저렴', '주차', '인증비', 
        '시험비', '처리'
    ],
    '기타': []  # 기본 카테고리
}

def categorize_opinion(opinion):
    """
    의견을 키워드 기반으로 카테고리 분류
    여러 카테고리에 해당할 수 있으므로 가장 관련성 높은 카테고리 반환
    """
    scores = {}
    
    for category, keywords in CATEGORIES.items():
        if category == '기타':
            continue
        score = sum(1 for keyword in keywords if keyword in opinion)
        if score > 0:
            scores[category] = score
    
    if not scores:
        return '기타'
    
    # 가장 높은 점수의 카테고리 반환
    return max(scores, key=scores.get)

def create_summary_sheet(wb, category_counts, total_count):
    """요약 시트 생성"""
    ws = wb.create_sheet('요약', 0)
    
    # 헤더
    ws['A1'] = '설문조사 주관식 의견 분석 요약'
    ws['A1'].font = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    # 전체 건수
    ws['A3'] = '전체 의견 건수'
    ws['B3'] = total_count
    ws['A3'].font = Font(name='맑은 고딕', size=11, bold=True)
    ws['B3'].font = Font(name='맑은 고딕', size=11)
    
    # 카테고리별 건수 헤더
    ws['A5'] = '카테고리'
    ws['B5'] = '건수'
    ws['C5'] = '비율(%)'
    ws['D5'] = '비율 그래프'
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}5'].font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        ws[f'{col}5'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 카테고리별 데이터
    row = 6
    for category in sorted(category_counts.keys(), key=lambda x: category_counts[x], reverse=True):
        count = category_counts[category]
        percentage = (count / total_count * 100) if total_count > 0 else 0
        
        ws[f'A{row}'] = category
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f'{percentage:.1f}%'
        ws[f'D{row}'] = '■' * int(percentage / 5)  # 5%당 1개의 블록
        
        ws[f'A{row}'].font = Font(name='맑은 고딕', size=10)
        ws[f'B{row}'].font = Font(name='맑은 고딕', size=10)
        ws[f'C{row}'].font = Font(name='맑은 고딕', size=10)
        ws[f'D{row}'].font = Font(name='맑은 고딕', size=10, color='4472C4')
        
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30

def create_category_sheet(wb, category_name, opinions_df):
    """카테고리별 시트 생성"""
    ws = wb.create_sheet(category_name)
    
    # 헤더
    ws['A1'] = f'카테고리: {category_name}'
    ws['A1'].font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 25
    
    # 건수 정보
    ws['A2'] = f'총 {len(opinions_df)}건'
    ws['A2'].font = Font(name='맑은 고딕', size=10, italic=True)
    ws.merge_cells('A2:B2')
    
    # 컬럼 헤더
    ws['A4'] = '번호'
    ws['B4'] = '의견 내용'
    
    for col in ['A', 'B']:
        ws[f'{col}4'].font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        ws[f'{col}4'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 데이터 입력
    for idx, (_, row) in enumerate(opinions_df.iterrows(), start=1):
        row_num = idx + 4
        ws[f'A{row_num}'] = idx
        ws[f'B{row_num}'] = row['의견']
        
        ws[f'A{row_num}'].font = Font(name='맑은 고딕', size=10)
        ws[f'B{row_num}'].font = Font(name='맑은 고딕', size=10)
        
        ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        ws.row_dimensions[row_num].height = max(15, len(row['의견']) / 50 * 15)
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 100

def create_all_opinions_sheet(wb, df):
    """전체 의견 시트 생성 (카테고리 태그 포함)"""
    ws = wb.create_sheet('전체 의견')
    
    # 헤더
    ws['A1'] = '전체 의견 목록 (카테고리 분류 포함)'
    ws['A1'].font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    
    # 컬럼 헤더
    ws['A3'] = '번호'
    ws['B3'] = '카테고리'
    ws['C3'] = '의견 내용'
    
    for col in ['A', 'B', 'C']:
        ws[f'{col}3'].font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        ws[f'{col}3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 데이터 입력
    for idx, (_, row) in enumerate(df.iterrows(), start=1):
        row_num = idx + 3
        ws[f'A{row_num}'] = idx
        ws[f'B{row_num}'] = row['카테고리']
        ws[f'C{row_num}'] = row['의견']
        
        ws[f'A{row_num}'].font = Font(name='맑은 고딕', size=10)
        ws[f'B{row_num}'].font = Font(name='맑은 고딕', size=10, bold=True)
        ws[f'C{row_num}'].font = Font(name='맑은 고딕', size=10)
        
        ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 카테고리별 색상 구분
        category_colors = {
            '홍보 및 정보 제공': 'FFE699',
            '장비 관리 및 운영': 'F8CBAD',
            '사용 절차 및 편의성': 'C5E0B4',
            '장비 및 서비스 확충': 'B4C7E7',
            '전문 지원 및 교육': 'D9D2E9',
            '비용 및 지원 사업': 'F4B084',
            '기타': 'D9D9D9'
        }
        color = category_colors.get(row['카테고리'], 'FFFFFF')
        ws[f'B{row_num}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        ws.row_dimensions[row_num].height = max(15, len(row['의견']) / 60 * 15)
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 100

def generate_markdown_report(df, category_counts, output_path):
    """마크다운 분석 보고서 생성"""
    total_count = len(df)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('# 설문조사 주관식 의견 분석 보고서\n\n')
        f.write(f'**분석 일시**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')
        f.write(f'**전체 의견 건수**: {total_count}건\n\n')
        
        f.write('## 카테고리별 분석 요약\n\n')
        f.write('| 카테고리 | 건수 | 비율 |\n')
        f.write('|---------|------|------|\n')
        
        for category in sorted(category_counts.keys(), key=lambda x: category_counts[x], reverse=True):
            count = category_counts[category]
            percentage = (count / total_count * 100) if total_count > 0 else 0
            f.write(f'| {category} | {count}건 | {percentage:.1f}% |\n')
        
        f.write('\n---\n\n')
        
        # 카테고리별 주요 의견
        for category in sorted(category_counts.keys(), key=lambda x: category_counts[x], reverse=True):
            if category == '기타':
                continue
                
            f.write(f'## {category} ({category_counts[category]}건)\n\n')
            
            category_opinions = df[df['카테고리'] == category]['의견'].tolist()
            
            # 주요 의견 3개 표시
            for idx, opinion in enumerate(category_opinions[:3], start=1):
                f.write(f'{idx}. {opinion}\n\n')
            
            if len(category_opinions) > 3:
                f.write(f'*...외 {len(category_opinions) - 3}건*\n\n')
            
            f.write('---\n\n')

def main():
    """메인 실행 함수"""
    print('=' * 80)
    print('설문조사 주관식 의견 분석 스크립트')
    print('=' * 80)
    
    # 파일 경로 설정
    input_file = r'd:\git_rk\data\g1_기타의견.csv'
    
    # REPORT 폴더 생성
    report_dir = r'd:\git_rk\REPORT'
    os.makedirs(report_dir, exist_ok=True)
    
    # 타임스탬프
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_excel = os.path.join(report_dir, f'설문조사_주관식의견_분석_{timestamp}.xlsx')
    output_markdown = os.path.join(report_dir, f'설문조사_주관식의견_분석_보고서_{timestamp}.md')
    
    print(f'\n[1/5] CSV 파일 읽기: {input_file}')
    
    # CSV 파일 읽기 (헤더 없음)
    with open(input_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    # 데이터 정리 (줄바꿈 제거)
    opinions = [line.strip() for line in lines if line.strip()]
    
    print(f'   총 {len(opinions)}건의 의견 로드 완료')
    
    print('\n[2/5] 카테고리 분류 진행 중...')
    
    # 카테고리 분류
    categorized_data = []
    for opinion in opinions:
        category = categorize_opinion(opinion)
        categorized_data.append({'의견': opinion, '카테고리': category})
    
    df = pd.DataFrame(categorized_data)
    
    # 카테고리별 건수 집계
    category_counts = df['카테고리'].value_counts().to_dict()
    
    print('   카테고리 분류 완료:')
    for category, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
        print(f'     - {category}: {count}건')
    
    print('\n[3/5] Excel 보고서 생성 중...')
    
    # Excel 파일 생성
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 요약 시트 생성
    create_summary_sheet(wb, category_counts, len(df))
    print('   요약 시트 생성 완료')
    
    # 카테고리별 시트 생성
    for category in sorted(category_counts.keys(), key=lambda x: category_counts[x], reverse=True):
        category_df = df[df['카테고리'] == category]
        create_category_sheet(wb, category, category_df)
        print(f'   "{category}" 시트 생성 완료')
    
    # 전체 의견 시트 생성
    create_all_opinions_sheet(wb, df)
    print('   전체 의견 시트 생성 완료')
    
    # Excel 파일 저장
    wb.save(output_excel)
    print(f'\n   Excel 파일 저장: {output_excel}')
    
    print('\n[4/5] 마크다운 보고서 생성 중...')
    
    # 마크다운 보고서 생성
    generate_markdown_report(df, category_counts, output_markdown)
    print(f'   마크다운 파일 저장: {output_markdown}')
    
    print('\n[5/5] 분석 완료!')
    print('=' * 80)
    print(f'\n생성된 파일:')
    print(f'  1. {output_excel}')
    print(f'  2. {output_markdown}')
    print('\n' + '=' * 80)

if __name__ == '__main__':
    main()
