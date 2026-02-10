# -*- coding: utf-8 -*-
"""csv_tables 폴더의 CSV → 단일 Excel 통합"""
import os, csv, sys
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

CSV_DIR = r'd:\git_rk\project\25_124_interviiew\260209\REPORT\output\csv_tables'
OUT_DIR = r'd:\git_rk\project\25_124_interviiew\260209\REPORT\output'
OUT_FILE = os.path.join(OUT_DIR, '분석결과_전체표.xlsx')

# 시트 이름 매핑 (파일명 → 시트명, 31자 제한)
def make_sheet_name(fname):
    name = fname.replace('.csv', '')
    # 31자 제한 처리
    if len(name) > 31:
        name = name[:31]
    return name

# 스타일 설정
header_font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='맑은 고딕', size=10)
data_align = Alignment(vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

wb = openpyxl.Workbook()
wb.remove(wb.active)  # 기본 시트 제거

# 시트 순서 정의
order_prefix = ['Z_조사', 'Z_규모별_특성', 'Z_업종별', 'Z_규모별',
                'A_', 'B_', 'C_', 'D_', 'E_']

csv_files = sorted(os.listdir(CSV_DIR))

# 정렬: Z → A → B → C → D → E 순
def sort_key(fname):
    for i, prefix in enumerate(order_prefix):
        if fname.startswith(prefix):
            return (i, fname)
    return (len(order_prefix), fname)

csv_files.sort(key=sort_key)

for fname in csv_files:
    if not fname.endswith('.csv'):
        continue
    
    fpath = os.path.join(CSV_DIR, fname)
    sheet_name = make_sheet_name(fname)
    
    # 중복 시트명 방지
    if sheet_name in wb.sheetnames:
        sheet_name = sheet_name[:28] + '_2'
    
    ws = wb.create_sheet(title=sheet_name)
    
    with open(fpath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = [r for r in reader if any(cell.strip() for cell in r)]
    
    if not rows:
        continue
    
    # markdown separator 행 제거 (:---: 패턴)
    filtered_rows = []
    for r in rows:
        if r and r[0].strip().startswith(':'):
            continue
        filtered_rows.append(r)
    rows = filtered_rows
    
    # 데이터 쓰기
    for ri, row in enumerate(rows):
        for ci, cell_val in enumerate(row):
            cell = ws.cell(row=ri+1, column=ci+1, value=cell_val.strip())
            cell.border = thin_border
            
            if ri == 0:  # 헤더
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            else:
                cell.font = data_font
                # NO 컬럼 등 숫자/짧은 텍스트는 가운데
                if ci == 0 or (cell_val.strip().isdigit()):
                    cell.alignment = center_align
                else:
                    cell.alignment = data_align
                # 짝수행 배경
                if ri % 2 == 0:
                    cell.fill = alt_fill
    
    # 컬럼 너비 자동 조절
    for ci in range(1, len(rows[0]) + 1):
        max_len = 0
        for ri in range(len(rows)):
            if ci <= len(rows[ri]):
                val = rows[ri][ci-1].strip()
                # 한글은 2배 폭
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, char_len)
        # 최소 8, 최대 50
        width = min(max(max_len + 2, 8), 50)
        ws.column_dimensions[get_column_letter(ci)].width = width
    
    # 첫 행 고정
    ws.freeze_panes = 'A2'
    
    print(f'[OK] {fname} -> {sheet_name}')

# 저장
wb.save(OUT_FILE)
print(f'\n[DONE] {OUT_FILE}')
print(f'총 시트 수: {len(wb.sheetnames)}')
