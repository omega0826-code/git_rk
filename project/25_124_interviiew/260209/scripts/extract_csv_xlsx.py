# -*- coding: utf-8 -*-
"""
통합보고서(00_integrated_report.md)에서 통계표를 CSV로 재추출하고
엑셀 취합 파일을 생성하는 스크립트
"""
import re, csv, os, sys

# openpyxl 사용
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORT = os.path.join(BASE, 'REPORT')
SRC = os.path.join(REPORT, '00_integrated_report.md')
OUT_CSV = os.path.join(REPORT, 'output', 'csv_tables')
OUT_XLSX = os.path.join(REPORT, 'output', '분석결과_전체표.xlsx')

print(f"[INFO] Source: {SRC}", flush=True)
print(f"[INFO] CSV output: {OUT_CSV}", flush=True)

# ── 1. 통합보고서 읽기 ──
with open(SRC, 'r', encoding='utf-8') as f:
    lines = f.readlines()

# ── 2. 마크다운 테이블 파싱 ──
def parse_md_tables(lines):
    """마크다운 라인에서 모든 테이블을 추출. 각 테이블은 (title, header_lines, data_lines)"""
    tables = []
    i = 0
    n = len(lines)
    current_section = ""
    current_subsection = ""
    
    while i < n:
        line = lines[i].rstrip()
        # 섹션 제목 추적
        if line.startswith('## '):
            current_section = line.lstrip('# ').strip()
        elif line.startswith('### '):
            current_subsection = line.lstrip('# ').strip()
        elif line.startswith('#### ') and '업종별' in line or line.startswith('#### ') and '규모별' in line:
            current_subsection = line.lstrip('# ').strip()
        
        # 테이블 시작 감지
        if '|' in line and line.strip().startswith('|'):
            table_lines = []
            title = current_subsection if current_subsection else current_section
            while i < n and '|' in lines[i] and lines[i].strip().startswith('|'):
                table_lines.append(lines[i].rstrip())
                i += 1
            
            # 구분선(---)만 있는 줄 제거하고, 최소 3줄(헤더+구분+데이터) 이상이면 테이블
            if len(table_lines) >= 3:
                tables.append((title, table_lines))
            continue
        i += 1
    
    return tables

def md_table_to_rows(table_lines):
    """마크다운 테이블 라인들을 2D 리스트로 변환"""
    rows = []
    for line in table_lines:
        cells = [c.strip() for c in line.split('|')]
        # 앞뒤 빈 셀 제거 (|로 시작하고 끝나므로)
        if cells and cells[0] == '':
            cells = cells[1:]
        if cells and cells[-1] == '':
            cells = cells[:-1]
        rows.append(cells)
    return rows

def is_separator_row(row):
    """구분선 행인지 확인"""
    return all(re.match(r'^[-:]+$', cell.strip()) or cell.strip() == '' for cell in row)

# ── 3. 테이블 추출 및 CSV 저장 ──
tables = parse_md_tables(lines)
print(f"[INFO] Found {len(tables)} tables", flush=True)

# 업종별 통계표만 필터 (기업별 응답 테이블 제외 - NO 컬럼 포함)
stat_tables = []
company_tables = []

for title, tlines in tables:
    rows = md_table_to_rows(tlines)
    if len(rows) < 2:
        continue
    
    # 구분선 제거
    data_rows = [r for r in rows if not is_separator_row(r)]
    if len(data_rows) < 2:
        continue
    
    # 첫 행에 'NO' 있으면 기업별 응답
    header = data_rows[0]
    if any('NO' in cell.upper() for cell in header[:2]):
        company_tables.append((title, data_rows))
    else:
        stat_tables.append((title, data_rows))

print(f"[INFO] Stat tables: {len(stat_tables)}, Company tables: {len(company_tables)}", flush=True)

# 파트별 매핑
part_map = {
    'A-1': 'A', 'A-2': 'A',
    'B-1': 'B', 'B-2': 'B', 'B-3': 'B', 'B-4': 'B',
    'C-1': 'C', 'C-2': 'C', 'C-3': 'C',
    'D-1': 'D', 'D-2': 'D', 'D-3': 'D',
    'E-1': 'E', 'E-2': 'E', 'E-3': 'E', 'E-4': 'E', 'E-5': 'E',
}

def get_part_prefix(title):
    for key, val in part_map.items():
        if key in title:
            return val
    if '종합' in title or '업종별 특성' in title or '규모별 특성' in title:
        return 'Z'
    return 'X'

def sanitize_filename(name, max_len=50):
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = re.sub(r'\s+', '_', name)
    return name[:max_len]

# CSV 저장
os.makedirs(OUT_CSV, exist_ok=True)
csv_files_created = []

# 통계표 CSV
for idx, (title, rows) in enumerate(stat_tables):
    part = get_part_prefix(title)
    fname = f"{part}_{sanitize_filename(title)}.csv"
    fpath = os.path.join(OUT_CSV, fname)
    
    with open(fpath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)
    
    csv_files_created.append((fname, 'stat', title))
    print(f"  [CSV] {fname}", flush=True)

# 기업별 응답 테이블 CSV
for idx, (title, rows) in enumerate(company_tables):
    part = get_part_prefix(title)
    fname = f"{part}_{sanitize_filename(title)}_기업별.csv"
    fpath = os.path.join(OUT_CSV, fname)
    
    with open(fpath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)
    
    csv_files_created.append((fname, 'company', title))
    print(f"  [CSV] {fname}", flush=True)

print(f"\n[INFO] Total CSV files: {len(csv_files_created)}", flush=True)

# ── 4. 엑셀 취합 파일 생성 ──
if not HAS_OPENPYXL:
    print("[WARN] openpyxl not installed, skipping Excel generation", flush=True)
    print("[INFO] Install: pip install openpyxl", flush=True)
    sys.exit(0)

wb = Workbook()
# 기본 시트 제거
wb.remove(wb.active)

# 스타일 정의
header_font = Font(name='맑은 고딕', bold=True, size=10)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
data_font = Font(name='맑은 고딕', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_table_to_sheet(ws, rows, start_row=1, title=None):
    """테이블 데이터를 시트에 작성"""
    r = start_row
    
    if title:
        ws.cell(row=r, column=1, value=title).font = Font(name='맑은 고딕', bold=True, size=12)
        r += 1
    
    for row_idx, row_data in enumerate(rows):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=r, column=col_idx + 1, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if row_idx == 0:  # 헤더
                cell.font = header_font_white
                cell.fill = header_fill
            else:
                cell.font = data_font
                # 숫자면 숫자로 변환
                try:
                    cell.value = int(value)
                except (ValueError, TypeError):
                    pass
        r += 1
    
    return r

# 파트별로 시트 생성
part_names = {
    'A': 'A_교육운영',
    'B': 'B_반복학습',
    'C': 'C_교육콘텐츠',
    'D': 'D_교육방식',
    'E': 'E_추가의견',
    'Z': 'Z_종합분석',
    'X': '기타',
}

# 통계표를 파트별로 그룹화
from collections import defaultdict
part_tables = defaultdict(list)

for fname, ttype, title in csv_files_created:
    part = fname[0]
    part_tables[part].append((fname, ttype, title))

for part in ['A', 'B', 'C', 'D', 'E', 'Z', 'X']:
    if part not in part_tables:
        continue
    
    sheet_name = part_names.get(part, part)
    ws = wb.create_sheet(title=sheet_name)
    
    current_row = 1
    
    for fname, ttype, title in part_tables[part]:
        fpath = os.path.join(OUT_CSV, fname)
        
        with open(fpath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        current_row = write_table_to_sheet(ws, rows, current_row, title)
        current_row += 1  # 테이블 간 빈 줄
    
    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 8), 30)

# 저장
wb.save(OUT_XLSX)
print(f"\n[SUCCESS] Excel: {OUT_XLSX}", flush=True)
print(f"[SUCCESS] CSV files: {len(csv_files_created)} files in {OUT_CSV}", flush=True)
print("[DONE]", flush=True)
