# -*- coding: utf-8 -*-
"""CSV -> Excel consolidation. Requires openpyxl."""
import csv, os, glob, sys
from datetime import datetime
print("START", flush=True)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BASE = r"d:\git_rk\project\25_124_interviiew\260209\REPORT\output"
CSV_DIR = os.path.join(BASE, "csv_tables")
TIMESTAMP = datetime.now().strftime("%y%m%d_%H%M")
OUT = os.path.join(BASE, f"result_all_tables_{TIMESTAMP}.xlsx")

wb = Workbook()
wb.remove(wb.active)

hdr_font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="4472C4")
dat_font = Font(name="맑은 고딕", size=10)
title_font = Font(name="맑은 고딕", bold=True, size=11)
bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

csvs = sorted(glob.glob(os.path.join(CSV_DIR, "*.csv")))
print(f"CSV files found: {len(csvs)}", flush=True)

# Group by first letter of filename
groups = {}
for fp in csvs:
    fn = os.path.basename(fp)
    key = fn[0] if fn[0].isalpha() else "X"
    groups.setdefault(key, []).append(fp)

sheet_names = {
    "A": "A_교육운영",
    "B": "B_반복학습", 
    "C": "C_교육콘텐츠",
    "D": "D_교육방식",
    "E": "E_추가의견",
    "Z": "Z_종합",
}

for key in sorted(groups.keys()):
    sn = sheet_names.get(key, key)
    ws = wb.create_sheet(sn)
    r = 1
    
    for fp in groups[key]:
        fn = os.path.basename(fp)
        title = fn[2:].replace(".csv", "").replace("_", " ")
        
        ws.cell(r, 1, title).font = title_font
        r += 1
        
        with open(fp, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for ri, row in enumerate(reader):
                for ci, v in enumerate(row):
                    cell = ws.cell(r, ci + 1)
                    cell.border = bdr
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    if ri == 0:
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                    else:
                        cell.font = dat_font
                    
                    # Try int conversion
                    try:
                        cell.value = int(v)
                    except (ValueError, TypeError):
                        cell.value = v
                r += 1
        r += 1  # blank row

    # Auto-width
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(mx + 2, 8), 35)

wb.save(OUT)
print(f"SAVED: {OUT}", flush=True)

# Also copy as Korean name
import shutil
OUT_KR = os.path.join(BASE, f"분석결과_전체표_{TIMESTAMP}.xlsx")
shutil.copy2(OUT, OUT_KR)
print(f"COPY: {OUT_KR}", flush=True)
print("DONE", flush=True)
