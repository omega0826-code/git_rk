"""
울산지역 기업체 교육훈련 인터뷰 데이터 엑셀 변환 스크립트 (분석 요약 포함)
작성일: 2026-01-21

각 시트 구성:
1. 분석 요약 섹션 (상단)
2. 원본 응답 데이터 (하단)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys

def main():
    print("=" * 80)
    print("인터뷰 데이터 엑셀 변환 시작 (분석 요약 포함)")
    print("=" * 80)
    
    # 파일 경로 설정
    csv_file = r'd:\git_rk\project\25_121_ulsan\25_124_interviiew\data\(CSV)interview_data_260121.csv'
    output_file = r'd:\git_rk\project\25_121_ulsan\25_124_interviiew\REPORT\인터뷰데이터_문항별정리_260121_0957.xlsx'
    
    # CSV 파일 읽기
    print(f"\n[1단계] CSV 파일 읽기")
    print(f"  파일: {csv_file}")
    
    try:
        df = pd.read_csv(csv_file, encoding='utf-8-sig')
        print(f"  ✓ 성공: {len(df)}개 기업 데이터 로드")
    except Exception as e:
        print(f"  ✗ 오류: {e}")
        return False
    
    # 문항별 분석 요약 데이터
    analysis_summary = {
        'Q1': {
            'title': '선호하는 교육 요일과 시간대',
            'summary': [
                ['선호 요일', '응답 기업 수', '비율'],
                ['화~목 평일', '12개사', '57%'],
                ['요일 무관', '5개사', '24%'],
                ['수요일 선호', '4개사', '19%'],
                ['', '', ''],
                ['선호 시간대', '응답 기업 수', '비율'],
                ['오후 (13:00~17:00)', '14개사', '67%'],
                ['시간 무관', '4개사', '19%'],
                ['근무시간 외', '3개사', '14%'],
                ['', '', ''],
                ['교육 어려운 시기', '응답 기업 수', '비율'],
                ['연초·연말', '15개사', '71%'],
                ['월초·월말', '3개사', '14%'],
                ['여름(7~8월)', '2개사', '10%'],
            ]
        },
        'Q2': {
            'title': '재직자 교육 추진 과정의 애로사항',
            'summary': [
                ['애로사항 유형', '응답 기업 수', '비율'],
                ['시간 확보 어려움', '16개사', '76%'],
                ['교육 콘텐츠 부족', '8개사', '38%'],
                ['참여도 저조', '7개사', '33%'],
                ['교육비 부담', '3개사', '14%'],
                ['접근성 문제', '2개사', '10%'],
                ['', '', ''],
                ['필요 지원사항', '응답 기업 수', '비율'],
                ['산업별 맞춤형 교육', '9개사', '43%'],
                ['유연한 교육 시간', '8개사', '38%'],
                ['실무 중심 교육', '7개사', '33%'],
                ['인센티브 제도', '5개사', '24%'],
                ['온라인 교육 확대', '4개사', '19%'],
            ]
        },
        'Q3': {
            'title': '동일 주제 교육 반복 수강 여부',
            'summary': [
                ['응답 유형', '기업 수', '비율'],
                ['반복 수강함', '12개사', '57%'],
                ['반복 수강 안 함', '6개사', '29%'],
                ['무응답/해당없음', '3개사', '14%'],
                ['', '', ''],
                ['반복 수강 이유', '응답 기업 수', ''],
                ['법규·규정 변경', '8개사', ''],
                ['이해도 향상', '5개사', ''],
                ['강사별 다른 관점', '4개사', ''],
                ['의무 교육', '3개사', ''],
            ]
        },
        'Q3_1': {
            'title': '강사 변경이 수강 의향에 미치는 영향',
            'summary': [
                ['응답', '기업 수', '비율'],
                ['영향 있음 (긍정적)', '6개사', '29%'],
                ['영향 없음', '11개사', '52%'],
                ['무응답', '4개사', '19%'],
            ]
        },
        'Q3_2': {
            'title': '최신 기술 동향 업데이트 시 재수강 의향',
            'summary': [
                ['응답', '기업 수', '비율'],
                ['재수강 의향 높음', '17개사', '81%'],
                ['상관없음', '2개사', '10%'],
                ['무응답', '2개사', '9%'],
                ['', '', ''],
                ['관심 분야', '응답 기업 수', ''],
                ['AI·디지털', '15개사', ''],
                ['법규·규제', '12개사', ''],
                ['ESG 경영', '4개사', ''],
                ['자동화·디지털 트윈', '3개사', ''],
            ]
        },
        'Q4': {
            'title': '반복 수강이 필요한 업무 영역',
            'summary': [
                ['주제', '응답 기업 수', '비율'],
                ['안전·보건 교육', '16개사', '76%'],
                ['인사·노무', '12개사', '57%'],
                ['법규·환경 규제', '8개사', '38%'],
                ['세무·회계', '6개사', '29%'],
                ['AI·디지털', '5개사', '24%'],
                ['ESG 경영', '3개사', '14%'],
                ['품질관리', '3개사', '14%'],
            ]
        },
        'Q5': {
            'title': '교육 효과를 높이기 위한 필요 지원',
            'summary': [
                ['지원 유형', '응답 기업 수', '비율'],
                ['짧은 압축형 교육', '8개사', '38%'],
                ['직무 맞춤형 콘텐츠', '7개사', '33%'],
                ['실무 중심 교육', '6개사', '29%'],
                ['교육자료 제공', '4개사', '19%'],
                ['인센티브 제도', '4개사', '19%'],
                ['난이도 조절', '2개사', '10%'],
                ['참여형 교육', '1개사', '5%'],
            ]
        },
        'Q6': {
            'title': 'AI 적용 업무 및 교육 방향',
            'summary': [
                ['AI 적용 현황', '기업 수', '비율'],
                ['적용 계획 있음', '9개사', '43%'],
                ['현재 없음 (향후 검토)', '8개사', '38%'],
                ['개인별 활용 중', '4개사', '19%'],
                ['', '', ''],
                ['AI 적용 분야', '응답 기업 수', ''],
                ['데이터 분석·예측', '6개사', ''],
                ['문서 자동화', '5개사', ''],
                ['품질관리·이상탐지', '4개사', ''],
                ['공정 최적화', '3개사', ''],
                ['업무 자동화', '3개사', ''],
                ['설계·CAD', '2개사', ''],
            ]
        },
        'Q7': {
            'title': '평일 온라인 + 주말 집체 교육 참여 가능성',
            'summary': [
                ['참여 가능 여부', '기업 수', '비율'],
                ['가능 (조건부 포함)', '10개사', '48%'],
                ['어려움', '8개사', '38%'],
                ['무응답', '3개사', '14%'],
                ['', '', ''],
                ['참여 가능 과목', '응답 기업 수', ''],
                ['인사·노무·세무', '8개사', ''],
                ['안전 교육', '6개사', ''],
                ['AI·디지털', '5개사', ''],
                ['CAD·설계', '3개사', ''],
                ['ERP·회계', '3개사', ''],
                ['ESG 경영', '2개사', ''],
            ]
        },
        'Q7_1': {
            'title': '평일 온라인 + 주말 집체 방식의 선호·우려사항',
            'summary': [
                ['선호사항', '응답 기업 수', ''],
                ['평일 온라인 접근성', '6개사', ''],
                ['주말 집체 실습·토론 가능', '5개사', ''],
                ['시간 제약 적음', '4개사', ''],
                ['', '', ''],
                ['우려사항', '응답 기업 수', '비율'],
                ['주말 참여도 저조', '12개사', '57%'],
                ['직원 피로도 증가', '5개사', '24%'],
                ['보상 문제', '3개사', '14%'],
                ['교대근무자 배려 필요', '2개사', '10%'],
            ]
        },
        'Q7_2': {
            'title': '대안적 교육 방식',
            'summary': [
                ['교육 방식', '응답 기업 수', '비율'],
                ['100% 온라인', '9개사', '43%'],
                ['평일 집체', '4개사', '19%'],
                ['하이브리드 (이론 온라인 + 실습 집체)', '3개사', '14%'],
                ['야간 교육', '2개사', '10%'],
                ['선택적 운영 (주중/주말 선택)', '2개사', '10%'],
                ['무응답', '1개사', '4%'],
            ]
        },
        'Q8': {
            'title': 'AI·디지털 교육 확대 시 필수 포함 내용',
            'summary': [
                ['교육 내용', '응답 기업 수', ''],
                ['ChatGPT·생성형 AI 실무 활용', '12개사', ''],
                ['산업별 AI 사례', '8개사', ''],
                ['AI 기초 교육', '6개사', ''],
                ['데이터 분석·활용', '5개사', ''],
                ['코딩 기초', '4개사', ''],
                ['디지털 전환', '3개사', ''],
                ['다양한 AI 툴', '3개사', ''],
                ['프롬프트 엔지니어링', '2개사', ''],
            ]
        },
        'Q_ETC': {
            'title': '기타 의견',
            'summary': [
                ['주요 의견', '기업', ''],
                ['교육 일정 사전 공지', '아성정밀, 대송컨테이너항만', ''],
                ['주차장 확보', '아성정밀, 현대미숀', ''],
                ['최신 AI 교육 필요', '디어벨리', ''],
                ['설문조사 창구 일원화', '㈜아일', ''],
                ['식사 제공', '영남파워', ''],
                ['비즈니스 회화 교육', '케이에이알', ''],
                ['북구 근처 교육기관', '세영윈도우', ''],
                ['신규직원 필수 교육', '수성정밀', ''],
            ]
        }
    }
    
    # 엑셀 파일 생성
    print(f"\n[2단계] 엑셀 파일 생성")
    print(f"  파일: {output_file}")
    
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            sheet_count = 0
            
            for q_code in ['Q1', 'Q2', 'Q3', 'Q3_1', 'Q3_2', 'Q4', 'Q5', 'Q6', 'Q7', 'Q7_1', 'Q7_2', 'Q8', 'Q_ETC']:
                if q_code in df.columns:
                    print(f"  - {q_code} 시트 생성 중...", end='')
                    
                    # 원본 응답 데이터 준비
                    response_data = []
                    for idx, row in df.iterrows():
                        company = row['업체명']
                        answer = row[q_code]
                        
                        if pd.notna(answer) and str(answer).strip() != '':
                            response_data.append({
                                '순번': idx + 1,
                                '기업명': company,
                                '응답내용(원본)': answer
                            })
                    
                    # 빈 시트 생성
                    pd.DataFrame().to_excel(writer, sheet_name=q_code, index=False)
                    worksheet = writer.sheets[q_code]
                    
                    current_row = 1
                    
                    # 1. 문항 제목
                    worksheet[f'A{current_row}'] = f"[{q_code}] {analysis_summary[q_code]['title']}"
                    worksheet.merge_cells(f'A{current_row}:C{current_row}')
                    title_cell = worksheet[f'A{current_row}']
                    title_cell.font = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
                    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    title_cell.alignment = Alignment(horizontal='left', vertical='center')
                    worksheet.row_dimensions[current_row].height = 30
                    current_row += 1
                    
                    # 2. 분석 요약 섹션
                    worksheet[f'A{current_row}'] = '📊 분석 요약'
                    worksheet.merge_cells(f'A{current_row}:C{current_row}')
                    summary_title = worksheet[f'A{current_row}']
                    summary_title.font = Font(name='맑은 고딕', size=11, bold=True)
                    summary_title.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    summary_title.alignment = Alignment(horizontal='left', vertical='center')
                    worksheet.row_dimensions[current_row].height = 25
                    current_row += 1
                    
                    # 분석 요약 데이터 작성
                    thin_border = Border(
                        left=Side(style='thin', color='D0D0D0'),
                        right=Side(style='thin', color='D0D0D0'),
                        top=Side(style='thin', color='D0D0D0'),
                        bottom=Side(style='thin', color='D0D0D0')
                    )
                    
                    for row_data in analysis_summary[q_code]['summary']:
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                            cell.font = Font(name='맑은 고딕', size=10)
                            cell.border = thin_border
                            
                            # 헤더 행 스타일
                            if row_data == analysis_summary[q_code]['summary'][0] or (len(row_data) > 0 and row_data[0] in ['선호 요일', '선호 시간대', '교육 어려운 시기', '애로사항 유형', '필요 지원사항', '응답 유형', '반복 수강 이유', '응답', '관심 분야', 'AI 적용 현황', 'AI 적용 분야', '참여 가능 여부', '참여 가능 과목', '선호사항', '우려사항', '교육 방식', '교육 내용', '주요 의견']):
                                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                                cell.font = Font(name='맑은 고딕', size=10, bold=True)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
                        
                        current_row += 1
                    
                    current_row += 1  # 빈 행
                    
                    # 3. 원본 응답 섹션
                    worksheet[f'A{current_row}'] = '📝 원본 응답 데이터'
                    worksheet.merge_cells(f'A{current_row}:C{current_row}')
                    response_title = worksheet[f'A{current_row}']
                    response_title.font = Font(name='맑은 고딕', size=11, bold=True)
                    response_title.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    response_title.alignment = Alignment(horizontal='left', vertical='center')
                    worksheet.row_dimensions[current_row].height = 25
                    current_row += 1
                    
                    # 원본 응답 헤더
                    headers = ['순번', '기업명', '응답내용(원본)']
                    for col_idx, header in enumerate(headers, start=1):
                        cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                        cell.font = Font(name='맑은 고딕', size=10, bold=True)
                        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    current_row += 1
                    
                    # 원본 응답 데이터
                    for data in response_data:
                        worksheet.cell(row=current_row, column=1, value=data['순번'])
                        worksheet.cell(row=current_row, column=2, value=data['기업명'])
                        worksheet.cell(row=current_row, column=3, value=data['응답내용(원본)'])
                        
                        for col_idx in range(1, 4):
                            cell = worksheet.cell(row=current_row, column=col_idx)
                            cell.font = Font(name='맑은 고딕', size=10)
                            cell.border = thin_border
                            
                            if col_idx == 1:
                                cell.alignment = Alignment(horizontal='center', vertical='top')
                            elif col_idx == 2:
                                cell.alignment = Alignment(horizontal='center', vertical='top')
                            else:
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        
                        current_row += 1
                    
                    # 컬럼 너비 조정
                    worksheet.column_dimensions['A'].width = 8
                    worksheet.column_dimensions['B'].width = 20
                    worksheet.column_dimensions['C'].width = 100
                    
                    sheet_count += 1
                    print(f" 완료 (요약 + {len(response_data)}개 응답)")
        
        print(f"\n[3단계] 완료")
        print(f"  ✓ 생성된 시트 수: {sheet_count}개")
        print(f"  ✓ 저장 위치: {output_file}")
        print("\n" + "=" * 80)
        return True
        
    except Exception as e:
        print(f"\n  ✗ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = main()
    if success:
        print("\n✅ 엑셀 파일 생성 성공!")
    else:
        print("\n❌ 엑셀 파일 생성 실패!")
        sys.exit(1)
