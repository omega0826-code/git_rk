#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PDF에서 표를 추출하여 Excel 파일로 변환하는 개선된 스크립트
여러 라이브러리를 결합하여 정확도를 높입니다.
"""

import os
import sys

# 라이브러리 import 및 설치 확인
try:
    import pdfplumber
except ImportError:
    print("pdfplumber가 설치되지 않았습니다. 설치 중...")
    os.system("pip install pdfplumber")
    import pdfplumber

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 설치되지 않았습니다. 설치 중...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Camelot 시도 (선택적)
try:
    import camelot
    CAMELOT_AVAILABLE = True
    print("✓ Camelot 사용 가능")
except ImportError:
    CAMELOT_AVAILABLE = False
    print("ℹ Camelot 미설치 (선택사항)")

# Tabula 시도 (선택적)
try:
    import tabula
    TABULA_AVAILABLE = True
    print("✓ Tabula 사용 가능")
except ImportError:
    TABULA_AVAILABLE = False
    print("ℹ Tabula 미설치 (선택사항)")


def extract_with_pdfplumber(pdf_path, table_settings=None):
    """
    pdfplumber를 사용한 표 추출 (기본 방법)
    """
    print("\n[방법 1] pdfplumber로 표 추출 중...")
    
    if table_settings is None:
        # 더 민감한 표 감지 설정
        table_settings = {
            "vertical_strategy": "lines_strict",  # 수직선 감지 전략
            "horizontal_strategy": "lines_strict",  # 수평선 감지 전략
            "explicit_vertical_lines": [],
            "explicit_horizontal_lines": [],
            "snap_tolerance": 3,  # 선 스냅 허용 오차
            "join_tolerance": 3,  # 선 결합 허용 오차
            "edge_min_length": 3,  # 최소 선 길이
            "min_words_vertical": 3,  # 수직 방향 최소 단어 수
            "min_words_horizontal": 1,  # 수평 방향 최소 단어 수
            "intersection_tolerance": 3,
        }
    
    tables_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            print(f"  페이지 {page_num}/{len(pdf.pages)} 처리 중...")
            
            # 기본 설정으로 표 추출
            tables = page.extract_tables(table_settings=table_settings)
            
            if tables:
                for table_idx, table in enumerate(tables):
                    if table and len(table) > 0:
                        tables_data.append({
                            'page': page_num,
                            'table_index': table_idx,
                            'data': table,
                            'method': 'pdfplumber_strict'
                        })
            
            # 더 느슨한 설정으로 재시도 (추가 표 감지)
            loose_settings = {
                "vertical_strategy": "text",  # 텍스트 기반 수직선 감지
                "horizontal_strategy": "text",  # 텍스트 기반 수평선 감지
                "snap_tolerance": 5,
                "join_tolerance": 5,
                "edge_min_length": 1,
            }
            
            loose_tables = page.extract_tables(table_settings=loose_settings)
            
            if loose_tables:
                for table_idx, table in enumerate(loose_tables):
                    if table and len(table) > 0:
                        # 중복 확인 (이미 추출된 표와 비교)
                        is_duplicate = False
                        for existing in tables_data:
                            if (existing['page'] == page_num and 
                                existing['data'] == table):
                                is_duplicate = True
                                break
                        
                        if not is_duplicate:
                            tables_data.append({
                                'page': page_num,
                                'table_index': len([t for t in tables_data if t['page'] == page_num]),
                                'data': table,
                                'method': 'pdfplumber_loose'
                            })
    
    print(f"  → pdfplumber로 {len(tables_data)}개 표 추출 완료")
    return tables_data


def extract_with_camelot(pdf_path):
    """
    Camelot을 사용한 표 추출 (선택적, 더 정확함)
    """
    if not CAMELOT_AVAILABLE:
        return []
    
    print("\n[방법 2] Camelot으로 표 추출 중...")
    tables_data = []
    
    try:
        # lattice 방식 (선이 있는 표)
        tables_lattice = camelot.read_pdf(pdf_path, pages='all', flavor='lattice')
        print(f"  Lattice 방식: {len(tables_lattice)}개 표 발견")
        
        for idx, table in enumerate(tables_lattice):
            tables_data.append({
                'page': table.page,
                'table_index': idx,
                'data': table.df.values.tolist(),
                'method': 'camelot_lattice',
                'accuracy': table.parsing_report.get('accuracy', 0)
            })
        
        # stream 방식 (선이 없는 표)
        tables_stream = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
        print(f"  Stream 방식: {len(tables_stream)}개 표 발견")
        
        for idx, table in enumerate(tables_stream):
            tables_data.append({
                'page': table.page,
                'table_index': idx,
                'data': table.df.values.tolist(),
                'method': 'camelot_stream',
                'accuracy': table.parsing_report.get('accuracy', 0)
            })
        
        print(f"  → Camelot으로 {len(tables_data)}개 표 추출 완료")
    except Exception as e:
        print(f"  ⚠ Camelot 추출 중 오류: {e}")
    
    return tables_data


def extract_with_tabula(pdf_path):
    """
    Tabula를 사용한 표 추출 (선택적)
    """
    if not TABULA_AVAILABLE:
        return []
    
    print("\n[방법 3] Tabula로 표 추출 중...")
    tables_data = []
    
    try:
        # lattice 방식
        tables_lattice = tabula.read_pdf(pdf_path, pages='all', lattice=True, multiple_tables=True)
        print(f"  Lattice 방식: {len(tables_lattice)}개 표 발견")
        
        for idx, df in enumerate(tables_lattice):
            if not df.empty:
                tables_data.append({
                    'page': idx + 1,  # Tabula는 페이지 정보를 직접 제공하지 않음
                    'table_index': idx,
                    'data': df.values.tolist(),
                    'method': 'tabula_lattice'
                })
        
        # stream 방식
        tables_stream = tabula.read_pdf(pdf_path, pages='all', stream=True, multiple_tables=True)
        print(f"  Stream 방식: {len(tables_stream)}개 표 발견")
        
        for idx, df in enumerate(tables_stream):
            if not df.empty:
                tables_data.append({
                    'page': idx + 1,
                    'table_index': idx,
                    'data': df.values.tolist(),
                    'method': 'tabula_stream'
                })
        
        print(f"  → Tabula로 {len(tables_data)}개 표 추출 완료")
    except Exception as e:
        print(f"  ⚠ Tabula 추출 중 오류: {e}")
    
    return tables_data


def merge_and_deduplicate_tables(all_tables):
    """
    여러 방법으로 추출한 표를 병합하고 중복 제거
    """
    print("\n표 병합 및 중복 제거 중...")
    
    # 페이지별로 그룹화
    tables_by_page = {}
    for table in all_tables:
        page = table['page']
        if page not in tables_by_page:
            tables_by_page[page] = []
        tables_by_page[page].append(table)
    
    # 중복 제거 및 최적 선택
    final_tables = []
    
    for page, page_tables in sorted(tables_by_page.items()):
        # 같은 위치의 표 중 가장 좋은 것 선택
        unique_tables = []
        
        for table in page_tables:
            is_duplicate = False
            
            for existing in unique_tables:
                # 데이터가 매우 유사하면 중복으로 간주
                if tables_are_similar(table['data'], existing['data']):
                    is_duplicate = True
                    # 더 정확한 방법으로 추출된 것으로 교체
                    if get_method_priority(table['method']) > get_method_priority(existing['method']):
                        unique_tables.remove(existing)
                        unique_tables.append(table)
                    break
            
            if not is_duplicate:
                unique_tables.append(table)
        
        final_tables.extend(unique_tables)
    
    print(f"  → 최종 {len(final_tables)}개 표 선택")
    return final_tables


def tables_are_similar(table1, table2, threshold=0.8):
    """
    두 표가 유사한지 확인
    """
    if not table1 or not table2:
        return False
    
    # 행 수가 크게 다르면 다른 표
    if abs(len(table1) - len(table2)) > 2:
        return False
    
    # 첫 몇 행을 비교
    rows_to_compare = min(3, len(table1), len(table2))
    similar_rows = 0
    
    for i in range(rows_to_compare):
        if i < len(table1) and i < len(table2):
            row1_str = str(table1[i])
            row2_str = str(table2[i])
            
            if row1_str == row2_str:
                similar_rows += 1
    
    similarity = similar_rows / rows_to_compare if rows_to_compare > 0 else 0
    return similarity >= threshold


def get_method_priority(method):
    """
    추출 방법의 우선순위 반환 (높을수록 좋음)
    """
    priorities = {
        'camelot_lattice': 5,
        'camelot_stream': 4,
        'pdfplumber_strict': 3,
        'tabula_lattice': 2,
        'pdfplumber_loose': 1,
        'tabula_stream': 1,
    }
    return priorities.get(method, 0)


def save_to_excel(tables, output_path):
    """
    추출된 표를 Excel 파일로 저장
    """
    print(f"\nExcel 파일 생성 중: {output_path}")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # 페이지 순서대로 정렬
    tables_sorted = sorted(tables, key=lambda x: (x['page'], x['table_index']))
    
    for table_num, table_info in enumerate(tables_sorted, start=1):
        sheet_name = f"표{table_num}"
        ws = wb.create_sheet(title=sheet_name)
        
        # 표 정보 헤더
        method_name = table_info['method'].replace('_', ' ').title()
        header_text = f"표 {table_num} (페이지 {table_info['page']}, 방법: {method_name})"
        ws.append([header_text])
        ws.append([])
        
        # 헤더 스타일링
        table_data = table_info['data']
        max_cols = max(len(row) for row in table_data) if table_data else 1
        
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f'A1:{get_column_letter(max_cols)}1')
        
        # 표 데이터 추가
        for row_idx, row in enumerate(table_data):
            cleaned_row = [cell if cell is not None else "" for cell in row]
            ws.append(cleaned_row)
            
            # 첫 번째 데이터 행 스타일링
            if row_idx == 0:
                current_row = ws.max_row
                for col_idx in range(1, len(cleaned_row) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 열 너비 자동 조정
        for col_idx in range(1, max_cols + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"  ✓ {sheet_name} 저장 완료")
    
    wb.save(output_path)
    print(f"\n✅ Excel 파일 저장 완료: {output_path}")


def main():
    """
    메인 실행 함수
    """
    print("=" * 60)
    print("PDF 표 추출 도구 (개선 버전)")
    print("=" * 60)
    
    # 파일 경로 설정
    pdf_file = "[보고서] 2024년도 울산지역 인력 및 훈련 수요공급조사 분석_통계편.pdf"
    output_file = "울산지역_인력훈련조사_표추출_개선.xlsx"
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pdf_path = os.path.join(script_dir, pdf_file)
    output_path = os.path.join(script_dir, output_file)
    
    # 파일 존재 확인
    if not os.path.exists(pdf_path):
        print(f"❌ 오류: PDF 파일을 찾을 수 없습니다: {pdf_path}")
        return
    
    print(f"\nPDF 파일: {pdf_file}")
    print(f"출력 파일: {output_file}\n")
    
    # 여러 방법으로 표 추출
    all_tables = []
    
    # 1. pdfplumber (기본, 항상 사용)
    all_tables.extend(extract_with_pdfplumber(pdf_path))
    
    # 2. Camelot (선택적, 설치되어 있으면 사용)
    all_tables.extend(extract_with_camelot(pdf_path))
    
    # 3. Tabula (선택적, 설치되어 있으면 사용)
    all_tables.extend(extract_with_tabula(pdf_path))
    
    if not all_tables:
        print("\n⚠️ 추출된 표가 없습니다.")
        return
    
    print(f"\n총 {len(all_tables)}개의 표 후보 발견")
    
    # 중복 제거 및 최적 선택
    final_tables = merge_and_deduplicate_tables(all_tables)
    
    # Excel로 저장
    save_to_excel(final_tables, output_path)
    
    print("\n" + "=" * 60)
    print(f"✅ 완료! 총 {len(final_tables)}개의 표를 추출했습니다.")
    print("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        input("\n아무 키나 눌러 종료...")
