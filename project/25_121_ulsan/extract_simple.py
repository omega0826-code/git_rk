#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PDF에서 표를 추출하여 Excel 파일로 변환하는 간소화된 스크립트
"""

import os
import sys

print("=" * 60)
print("PDF 표 추출 시작")
print("=" * 60)

# 1. 라이브러리 import
print("\n[1/4] 라이브러리 로딩 중...")
try:
    import pdfplumber
    print("  ✓ pdfplumber 로드 완료")
except ImportError:
    print("  ✗ pdfplumber 설치 필요")
    print("  설치 중...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pdfplumber"])
    import pdfplumber
    print("  ✓ pdfplumber 설치 및 로드 완료")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    print("  ✓ openpyxl 로드 완료")
except ImportError:
    print("  ✗ openpyxl 설치 필요")
    print("  설치 중...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    print("  ✓ openpyxl 설치 및 로드 완료")

# 2. 파일 경로 설정
print("\n[2/4] 파일 경로 설정 중...")
pdf_file = "[보고서] 2024년도 울산지역 인력 및 훈련 수요공급조사 분석_통계편.pdf"
output_file = "울산지역_인력훈련조사_표추출_최종.xlsx"

script_dir = os.path.dirname(os.path.abspath(__file__))
pdf_path = os.path.join(script_dir, pdf_file)
output_path = os.path.join(script_dir, output_file)

print(f"  PDF: {pdf_file}")
print(f"  출력: {output_file}")

if not os.path.exists(pdf_path):
    print(f"\n❌ 오류: PDF 파일을 찾을 수 없습니다!")
    print(f"  경로: {pdf_path}")
    input("\n아무 키나 눌러 종료...")
    sys.exit(1)

print("  ✓ PDF 파일 확인 완료")

# 3. PDF에서 표 추출
print("\n[3/4] PDF에서 표 추출 중...")
all_tables = []

with pdfplumber.open(pdf_path) as pdf:
    total_pages = len(pdf.pages)
    print(f"  총 {total_pages} 페이지")
    
    for page_num, page in enumerate(pdf.pages, start=1):
        print(f"  페이지 {page_num}/{total_pages} 처리 중...", end=" ")
        
        # 방법 1: Strict 모드 (선이 명확한 표)
        strict_settings = {
            "vertical_strategy": "lines_strict",
            "horizontal_strategy": "lines_strict",
            "snap_tolerance": 3,
            "join_tolerance": 3,
        }
        
        tables_strict = page.extract_tables(table_settings=strict_settings)
        
        # 방법 2: Loose 모드 (선이 없는 표)
        loose_settings = {
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
            "snap_tolerance": 5,
            "join_tolerance": 5,
        }
        
        tables_loose = page.extract_tables(table_settings=loose_settings)
        
        # 결과 합치기
        page_tables = []
        
        if tables_strict:
            for table in tables_strict:
                if table and len(table) > 0:
                    page_tables.append(table)
        
        if tables_loose:
            for table in tables_loose:
                if table and len(table) > 0:
                    # 중복 확인
                    is_duplicate = False
                    for existing in page_tables:
                        if existing == table:
                            is_duplicate = True
                            break
                    if not is_duplicate:
                        page_tables.append(table)
        
        # 페이지 정보와 함께 저장
        for idx, table in enumerate(page_tables):
            all_tables.append({
                'page': page_num,
                'index': idx,
                'data': table
            })
        
        print(f"{len(page_tables)}개 표 발견")

print(f"\n  ✓ 총 {len(all_tables)}개 표 추출 완료")

if not all_tables:
    print("\n⚠️ 추출된 표가 없습니다.")
    input("\n아무 키나 눌러 종료...")
    sys.exit(0)

# 4. Excel 파일 생성
print(f"\n[4/4] Excel 파일 생성 중...")
wb = Workbook()
wb.remove(wb.active)

for table_num, table_info in enumerate(all_tables, start=1):
    sheet_name = f"표{table_num}"
    ws = wb.create_sheet(title=sheet_name)
    
    # 표 정보 헤더
    header_text = f"표 {table_num} (페이지 {table_info['page']})"
    ws.append([header_text])
    ws.append([])
    
    # 헤더 스타일링
    table_data = table_info['data']
    max_cols = max(len(row) for row in table_data) if table_data else 1
    
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f'A1:{get_column_letter(max_cols)}1')
    
    # 표 데이터 추가
    for row_idx, row in enumerate(table_data):
        # openpyxl에서 허용하지 않는 제어 문자 제거
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append("")
            else:
                # 문자열로 변환하고 제어 문자 제거
                cell_str = str(cell)
                # 제어 문자 제거 (0x00-0x1F, 0x7F-0x9F 범위, 단 탭/줄바꿈/캐리지리턴 제외)
                cleaned_cell = ''.join(char for char in cell_str 
                                      if ord(char) >= 32 or char in '\t\n\r')
                cleaned_row.append(cleaned_cell)
        
        ws.append(cleaned_row)
        
        # 첫 번째 데이터 행 스타일링
        if row_idx == 0:
            current_row = ws.max_row
            for col_idx in range(1, len(cleaned_row) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 열 너비 자동 조정
    for col_idx in range(1, max_cols + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row in ws[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"  ✓ {sheet_name} 생성 완료")

# Excel 파일 저장
wb.save(output_path)

print("\n" + "=" * 60)
print(f"✅ 완료!")
print(f"  총 {len(all_tables)}개의 표를 추출했습니다.")
print(f"  파일: {output_path}")
print("=" * 60)
print("\n작업이 완료되었습니다!")
