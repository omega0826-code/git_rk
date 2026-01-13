#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PDF에서 표를 추출하여 Excel 파일로 변환하는 스크립트
각 표는 별도의 시트로 저장됩니다.
"""

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def extract_tables_to_excel(pdf_path, output_excel_path):
    """
    PDF 파일에서 모든 표를 추출하여 Excel 파일로 저장
    
    Args:
        pdf_path: PDF 파일 경로
        output_excel_path: 출력할 Excel 파일 경로
    """
    print(f"PDF 파일 열기: {pdf_path}")
    
    # Excel 워크북 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    table_count = 0
    
    # PDF 파일 열기
    with pdfplumber.open(pdf_path) as pdf:
        print(f"총 페이지 수: {len(pdf.pages)}")
        
        # 각 페이지 순회
        for page_num, page in enumerate(pdf.pages, start=1):
            print(f"\n페이지 {page_num} 처리 중...")
            
            # 페이지에서 표 추출
            tables = page.extract_tables()
            
            if tables:
                print(f"  - {len(tables)}개의 표 발견")
                
                # 각 표를 별도 시트로 저장
                for table_idx, table in enumerate(tables):
                    table_count += 1
                    sheet_name = f"표{table_count}"
                    
                    print(f"  - {sheet_name} 추출 중... (페이지 {page_num}, 표 {table_idx + 1})")
                    
                    # 새 시트 생성
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 표 정보 헤더 추가
                    ws.append([f"표 {table_count} (페이지 {page_num})"])
                    ws.append([])  # 빈 행 추가
                    
                    # 표 정보 헤더 스타일링
                    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
                    ws.merge_cells('A1:' + get_column_letter(len(table[0]) if table else 1) + '1')
                    
                    # 표 데이터 추가
                    for row_idx, row in enumerate(table):
                        # None 값을 빈 문자열로 변환
                        cleaned_row = [cell if cell is not None else "" for cell in row]
                        ws.append(cleaned_row)
                        
                        # 첫 번째 데이터 행(헤더)에 스타일 적용
                        if row_idx == 0:
                            current_row = ws.max_row
                            for col_idx in range(1, len(cleaned_row) + 1):
                                cell = ws.cell(row=current_row, column=col_idx)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 열 너비 자동 조정
                    for col_idx, col in enumerate(ws.columns, start=1):
                        max_length = 0
                        column_letter = get_column_letter(col_idx)
                        
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)  # 최대 50으로 제한
                        ws.column_dimensions[column_letter].width = adjusted_width
            else:
                print(f"  - 표 없음")
    
    # Excel 파일 저장
    if table_count > 0:
        wb.save(output_excel_path)
        print(f"\n✅ 완료! 총 {table_count}개의 표를 추출했습니다.")
        print(f"📁 저장 위치: {output_excel_path}")
    else:
        print("\n⚠️ 추출된 표가 없습니다.")
    
    return table_count

if __name__ == "__main__":
    # 파일 경로 설정
    pdf_file = "[보고서] 2024년도 울산지역 인력 및 훈련 수요공급조사 분석_통계편.pdf"
    output_file = "울산지역_인력훈련조사_표추출.xlsx"
    
    # 현재 스크립트 디렉토리 기준으로 경로 설정
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pdf_path = os.path.join(script_dir, pdf_file)
    output_path = os.path.join(script_dir, output_file)
    
    # 파일 존재 확인
    if not os.path.exists(pdf_path):
        print(f"❌ 오류: PDF 파일을 찾을 수 없습니다: {pdf_path}")
        exit(1)
    
    # 표 추출 실행
    try:
        extract_tables_to_excel(pdf_path, output_path)
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
