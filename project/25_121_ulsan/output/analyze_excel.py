# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'd:\git_rk\project\25_121_ulsan\output\(OUTPUT)울산 일자리 미스매치_260205_1822.xlsx')

with open(r'd:\git_rk\project\25_121_ulsan\output\_analysis_result.txt', 'w', encoding='utf-8') as out:
    out.write(f"Sheet count: {len(wb.sheetnames)}\n")
    out.write(f"Sheet names: {wb.sheetnames}\n")
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        out.write(f"\n=== Sheet: {sname} ===\n")
        out.write(f"  Rows: {ws.max_row}, Cols: {ws.max_column}\n")
        out.write(f"  Merged cells: {[str(m) for m in ws.merged_cells.ranges]}\n")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), start=1):
            vals = []
            for c in row:
                vals.append(str(c.value) if c.value is not None else "")
            out.write(f"  R{row_idx}: {vals}\n")

print("Done")
