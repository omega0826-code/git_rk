# -*- coding: utf-8 -*-
"""
SPSS 출력 결과 Excel -> Markdown 변환 스크립트
구조: 병합 셀, 다중 헤더, 교차표 등 SPSS 특유의 형식을 정밀하게 처리
"""
import openpyxl
import re
import os

def load_sheet(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]  # Sheet1
    
    # 병합 셀 정보 수집
    merged = {}
    for m in ws.merged_cells.ranges:
        val = ws.cell(m.min_row, m.min_col).value
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != (m.min_row, m.min_col):
                    merged[(r, c)] = val  # 병합된 셀에 원래 값 복원
    
    data = []
    for row_idx in range(1, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            val = cell.value
            if val is None and (row_idx, col_idx) in merged:
                val = merged[(row_idx, col_idx)]
            row_data.append(val)
        data.append(row_data)
    
    return data

def is_empty_row(row):
    return all(v is None or str(v).strip() == '' for v in row)

def fmt_val(v):
    """값을 보기 좋게 포맷"""
    if v is None or str(v).strip() == '':
        return '-'
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        # 비율이면 퍼센트로 변환
        if 0 < abs(v) < 1.01:
            return f"{v*100:.1f}%"
        # 소수점이 긴 경우 반올림
        if abs(v) > 1:
            return f"{v:.2f}"
        return f"{v:.4f}"
    return str(v).strip()

def fmt_val_raw(v):
    """원시 값 포맷 (비율 변환 없이)"""
    if v is None or str(v).strip() == '':
        return '-'
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return f"{v:.2f}"
    return str(v).strip()

def identify_tables(data):
    """데이터에서 개별 테이블 블록을 식별"""
    tables = []
    i = 0
    while i < len(data):
        row = data[i]
        # 제목 행 찾기: A열에 텍스트가 있고, 나머지가 대부분 빈 행
        if row[0] is not None and str(row[0]).strip() != '':
            first_val = str(row[0]).strip()
            # 빈 행이 아니고 데이터행도 아닌, 테이블 제목인지 확인
            non_empty = sum(1 for v in row if v is not None and str(v).strip() != '')
            if non_empty <= 4 and len(first_val) > 3:
                # 타이틀 후보 발견
                table_start = i
                title = first_val
                
                # 테이블 끝 찾기 (다음 빈 행까지)
                j = i + 1
                while j < len(data) and not is_empty_row(data[j]):
                    j += 1
                
                tables.append({
                    'title': title,
                    'start': table_start,
                    'end': j,
                    'data': data[table_start:j]
                })
                i = j + 1
                continue
        i += 1
    return tables

def build_table1_general(data):
    """응답자 일반현황 테이블 (R1~R15)"""
    lines = []
    lines.append("## 1. 응답자 일반현황\n")
    lines.append("| 구 분 | | 응답자 수 | 비율 |")
    lines.append("|:---|:---|---:|---:|")
    
    for row in data[2:]:  # R3부터
        a, b, c, d = row[0], row[1], row[2], row[3]
        if is_empty_row(row):
            break
        a_str = fmt_val_raw(a) if a else ''
        b_str = fmt_val_raw(b) if b else ''
        c_str = fmt_val(c)
        d_val = d
        if isinstance(d_val, float) and 0 < abs(d_val) <= 1:
            d_str = f"{d_val*100:.1f}%"
        else:
            d_str = fmt_val(d_val)
        lines.append(f"| {a_str} | {b_str} | {c_str} | {d_str} |")
    
    lines.append("")
    return '\n'.join(lines)

def build_cross_table(table_data, table_num, title):
    """교차표(빈도/비율 쌍) 변환"""
    lines = []
    lines.append(f"## {table_num}. {title}\n")
    
    rows = table_data
    
    # 헤더 영역 파악: 보통 2~4행이 헤더
    # 카테고리 헤더 행 (예: 학교 취업지원센터, 온라인 채용사이트...)
    # 빈도/비율 행
    
    # rows[0] = 제목, rows[1] = '구 분' 행, rows[2] = 카테고리 헤더, rows[3] = 빈도/비율
    # rows[4:] = 데이터
    
    if len(rows) < 4:
        lines.append("(데이터 부족)\n")
        return '\n'.join(lines)
    
    # 카테고리 헤더 추출
    header_row = rows[2]  # 카테고리명이 있는 행
    subheader_row = rows[3]  # 빈도/비율 행
    
    # 유효한 카테고리 추출 (빈도/비율 쌍)
    categories = []
    col = 3  # D열부터 시작
    while col < len(header_row):
        cat_name = header_row[col]
        if cat_name is not None and str(cat_name).strip() != '' and str(cat_name).strip() != '빈도' and str(cat_name).strip() != '비율':
            categories.append({'name': str(cat_name).strip(), 'col': col})
        col += 1
    
    # 중복 제거 (병합 셀로 인해 같은 이름이 반복될 수 있음)
    seen = set()
    unique_cats = []
    for cat in categories:
        if cat['name'] not in seen:
            seen.add(cat['name'])
            unique_cats.append(cat)
    categories = unique_cats
    
    if not categories:
        # 카테고리가 없는 경우 raw 테이블로 출력
        return build_simple_table(table_data, table_num, title)
    
    # 마크다운 테이블 생성 - 빈도와 비율을 각각 별도 열로
    header_line = "| 구 분 | | 응답자 수 |"
    sep_line = "|:---|:---|---:|"
    for cat in categories:
        header_line += f" {cat['name']}(빈도) | {cat['name']}(비율) |"
        sep_line += " ---:| ---:|"
    
    lines.append(header_line)
    lines.append(sep_line)
    
    # 데이터 행
    for row in rows[4:]:
        if is_empty_row(row):
            break
        a_str = fmt_val_raw(row[0]) if row[0] else ''
        b_str = fmt_val_raw(row[1]) if row[1] else ''
        c_str = fmt_val(row[2])
        
        line = f"| {a_str} | {b_str} | {c_str} |"
        for idx, cat in enumerate(categories):
            freq_col = 3 + idx * 2
            rate_col = 3 + idx * 2 + 1
            freq = row[freq_col] if freq_col < len(row) else None
            rate = row[rate_col] if rate_col < len(row) else None
            
            freq_str = fmt_val(freq) if freq else '-'
            if isinstance(rate, float) and 0 < abs(rate) <= 1:
                rate_str = f"{rate*100:.1f}%"
            else:
                rate_str = fmt_val(rate) if rate else '-'
            
            line += f" {freq_str} | {rate_str} |"
        
        lines.append(line)
    
    lines.append("")
    return '\n'.join(lines)

def build_simple_table(table_data, table_num, title):
    """단순 테이블 (카테고리 없는 경우)"""
    lines = []
    lines.append(f"## {table_num}. {title}\n")
    
    rows = table_data
    if len(rows) < 2:
        return '\n'.join(lines)
    
    # 헤더 결정
    # 첫 번째 비어있지 않은 행을 헤더로
    header_idx = 0
    for hidx, r in enumerate(rows):
        non_empty = [v for v in r if v is not None and str(v).strip() != '']
        if len(non_empty) >= 3:
            header_idx = hidx
            break
    
    header = rows[header_idx]
    cols = []
    for v in header:
        if v is not None and str(v).strip() != '':
            cols.append(str(v).strip())
    
    if not cols:
        return '\n'.join(lines)
    
    header_line = "| " + " | ".join(cols) + " |"
    sep_line = "| " + " | ".join(["---:" if i > 0 else ":---" for i in range(len(cols))]) + " |"
    
    lines.append(header_line)
    lines.append(sep_line)
    
    for row in rows[header_idx+1:]:
        if is_empty_row(row):
            break
        vals = [fmt_val_raw(row[i]) if i < len(row) else '-' for i in range(len(cols))]
        lines.append("| " + " | ".join(vals) + " |")
    
    lines.append("")
    return '\n'.join(lines)

def build_freq_table(table_data, table_num, title):
    """업종/직군 빈도 테이블 (병합 헤더, 대학별 응답자 수)"""
    lines = []
    lines.append(f"## {table_num}. {title}\n")
    
    rows = table_data
    # 헤더: 보통 첫 2-3행이 헤더
    # B229: 응답자 수, 소속대학 하위에 대학명
    
    # 헤더 파악
    header1 = rows[1]  # 응답자 수, 소속 대학...
    header2 = rows[2]  # 울산대, 울산과학대...
    
    lines.append("| 구 분 | 전체 응답자 수 | 울산대학교 | 울산과학대학교 | 춘해보건대학교 | 울산과학기술원 |")
    lines.append("|:---|---:|---:|---:|---:|---:|")
    
    for row in rows[4:]:  # 데이터 시작
        if is_empty_row(row):
            break
        name = fmt_val_raw(row[0]) if row[0] else '-'
        total = fmt_val(row[1]) if row[1] else '-'
        u1 = fmt_val(row[2]) if row[2] else '-'
        u2 = fmt_val(row[3]) if row[3] else '-'
        u3 = fmt_val(row[4]) if row[4] else '-'
        u4 = fmt_val(row[5]) if row[5] else '-'
        lines.append(f"| {name} | {total} | {u1} | {u2} | {u3} | {u4} |")
    
    lines.append("")
    return '\n'.join(lines)

def build_perception_table(table_data, table_num, title, value_label):
    """인식 수준 테이블 (연봉 수준, 근무 환경 등)"""
    lines = []
    lines.append(f"## {table_num}. {title} ({value_label})\n")
    
    rows = table_data
    
    # 헤더: 연봉 수준, 근무 환경 등
    header_row = rows[1]  # R270/R287
    
    # 카테고리 추출
    categories = []
    for col_idx in range(2, len(header_row)):
        v = header_row[col_idx]
        if v is not None and str(v).strip() != '' and str(v).strip() != value_label:
            categories.append(str(v).strip())
    
    # 중복 제거
    seen = set()
    unique_cats = []
    for c in categories:
        if c not in seen:
            seen.add(c)
            unique_cats.append(c)
    categories = unique_cats
    
    header_line = "| 구 분 | |"
    sep_line = "|:---|:---|"
    for cat in categories:
        header_line += f" {cat} |"
        sep_line += " ---:|"
    
    lines.append(header_line)
    lines.append(sep_line)
    
    for row in rows[3:]:  # 데이터 시작 (헤더 3행 건너띔)
        if is_empty_row(row):
            break
        a_str = fmt_val_raw(row[0]) if row[0] else ''
        b_str = fmt_val_raw(row[1]) if row[1] else ''
        line = f"| {a_str} | {b_str} |"
        for idx in range(len(categories)):
            col_idx = 2 + idx
            val = row[col_idx] if col_idx < len(row) else None
            line += f" {fmt_val_raw(val)} |"
        lines.append(line)
    
    lines.append("")
    return '\n'.join(lines)

def convert_to_markdown(filepath):
    data = load_sheet(filepath)
    
    md_parts = []
    md_parts.append("# 울산 일자리 미스매치 분석 결과\n")
    md_parts.append("> SPSS 통계 분석 출력 결과를 마크다운 형식으로 변환한 자료입니다.\n")
    md_parts.append("---\n")
    
    # ========================================
    # 1. 응답자 일반현황 (R1~R15)
    # ========================================
    md_parts.append("## 1. 응답자 일반현황\n")
    md_parts.append("| 구 분 | 세부 항목 | 응답자 수 | 비율 |")
    md_parts.append("|:---|:---|---:|---:|")
    for i in range(2, 15):  # R3~R15
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = int(row[2]) if isinstance(row[2], float) else row[2]
        d = row[3]
        d_str = f"{d*100:.1f}%" if isinstance(d, float) and 0 < abs(d) <= 1 else fmt_val(d)
        md_parts.append(f"| {a} | {b} | {c} | {d_str} |")
    md_parts.append("\n---\n")
    
    # ========================================
    # 2. 일자리 정보 탐색 활용 경로(중복응답) (R17~R33)
    # ========================================
    cats2 = ['학교 취업지원센터/커리어센터', '온라인 채용사이트', '기업 홈페이지 직접 방문', 'SNS', '지인 소개/추천', '공공기관 취업지원센터', '취업 카페/커뮤니티', '기타']
    md_parts.append("## 2. 일자리 정보 탐색 활용 경로(중복응답)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats2:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(20, 33):  # R21~R33
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats2)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 3. 취업 준비 중 느끼는 어려움(중복응답) (R35~R51)
    # ========================================
    cats3 = ['직무 관련 역량 부족', '자격증/어학 점수 부족', '경험 및 경력 부족', '취업 정보 부족', '자기소개서/면접 준비 어려움', '경제적 부담', '심리적 스트레스', '기타']
    md_parts.append("## 3. 취업 준비 중 느끼는 어려움(중복응답)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats3:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(38, 51):  # R39~R51
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats3)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 4. 취업 준비 중 대학에서 제공 받았으면 하는 지원(중복응답) (R53~R69)
    # ========================================
    cats4 = ['취업 정보 제공', '취업 관련 교육', '자격증 취득', '취업 맨토링/컨설팅', '인턴십 기회 확대', '취업 연계 프로그램', '기타']
    md_parts.append("## 4. 취업 준비 중 대학에서 제공 받았으면 하는 지원(중복응답)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats4:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(56, 69):  # R57~R69
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats4)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 5. 취업 준비 중 대학에서 제공 받았으면 하는 지원(1순위) (R71~R87)
    # ========================================
    cats5 = ['채용공고 및 채용일정 정보', '기업별 연봉 및 복지 정보', '직무별 구체적인 업무 내용', '기업이 요구하는 자격요건 및 우대사항', '인턴십/현장실습 기회 정보', '합격자 스펙 및 면접 후기']
    md_parts.append("## 5. 취업 준비 중 대학에서 제공 받았으면 하는 지원(1순위)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats5:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(74, 87):  # R75~R87
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats5)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 6. 취업 준비 중 대학에서 제공 받았으면 하는 지원(1순위+2순위) (R89~R105)
    # ========================================
    cats6 = ['채용공고 및 채용일정 정보', '기업별 연봉 및 복지 정보', '직무별 구체적인 업무 내용', '기업이 요구하는 자격요건 및 우대사항', '인턴십/현장실습 기회 정보', '합격자 스펙 및 면접 후기', '기타']
    md_parts.append("## 6. 취업 준비 중 대학에서 제공 받았으면 하는 지원(1순위+2순위)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats6:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(92, 105):  # R93~R105
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats6)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 7. 취업 준비 중 본인에게 부족함을 느끼는 역량(1순위) (R107~R123)
    # ========================================
    cats7 = ['전공 지식', '외국어 능력', '컴퓨터 활용 능력', '의사소통 능력', '문제해결 능력', '대인관계 능력', '창의력/기획력', '리더십', '특별히 없음', '기타']
    md_parts.append("## 7. 취업 준비 중 본인에게 부족함을 느끼는 역량(1순위)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats7:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(110, 123):  # R111~R123
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats7)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 8. 취업 준비 중 본인에게 부족함을 느끼는 역량(1순위+2순위) (R125~R141)
    # ========================================
    md_parts.append("## 8. 취업 준비 중 본인에게 부족함을 느끼는 역량(1순위+2순위)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats7:  # 같은 카테고리
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(128, 141):  # R129~R141
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats7)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 9. 취업 시 중요하게 고려하는 요소(1순위) (R143~R159)
    # ========================================
    cats9 = ['연봉', '복지', '근무지', '기업 규모', '기업 인지도', '직무 적합성', '고용 안정성', '성장 가능성', '기타']
    md_parts.append("## 9. 취업 시 중요하게 고려하는 요소(1순위)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats9:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(146, 159):  # R147~R159
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats9)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 10. 취업 시 중요하게 고려하는 요소(1순위+2순위) (R161~R177)
    # ========================================
    md_parts.append("## 10. 취업 시 중요하게 고려하는 요소(1순위+2순위)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats9:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(164, 177):  # R165~R177
        row = data[i]
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats9)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 11. 대학교별 취업 희망 세부 업종 (R179~R226)
    # ========================================
    md_parts.append("## 11. 대학교별 취업 희망 세부 업종\n")
    md_parts.append("| 업종 | 전체 응답자 수 | 울산대학교 | 울산과학대학교 | 춘해보건대학교 | 울산과학기술원 |")
    md_parts.append("|:---|---:|---:|---:|---:|---:|")
    for i in range(182, 226):  # R183~R226
        row = data[i]
        if is_empty_row(row):
            break
        name = str(row[0]).strip() if row[0] else '-'
        total = fmt_val(row[1])
        u1 = fmt_val(row[2])
        u2 = fmt_val(row[3])
        u3 = fmt_val(row[4])
        u4 = fmt_val(row[5])
        md_parts.append(f"| {name} | {total} | {u1} | {u2} | {u3} | {u4} |")
    md_parts.append("\n---\n")
    
    # ========================================
    # 12. 대학교별 취업 희망 직군 (R228~R267)
    # ========================================
    md_parts.append("## 12. 대학교별 취업 희망 직군\n")
    md_parts.append("| 직군 | 전체 응답자 수 | 울산대학교 | 울산과학대학교 | 춘해보건대학교 | 울산과학기술원 |")
    md_parts.append("|:---|---:|---:|---:|---:|---:|")
    for i in range(231, 267):  # R232~R267
        row = data[i]
        if is_empty_row(row):
            break
        name = str(row[0]).strip() if row[0] else '-'
        total = fmt_val(row[1])
        u1 = fmt_val(row[2])
        u2 = fmt_val(row[3])
        u3 = fmt_val(row[4])
        u4 = fmt_val(row[5])
        md_parts.append(f"| {name} | {total} | {u1} | {u2} | {u3} | {u4} |")
    md_parts.append("\n---\n")
    
    # ========================================
    # 13. 울산 지역 기업 인식 수준 - 응답자 수 (R269~R284)
    # ========================================
    perception_cats = ['연봉 수준', '근무 환경', '물리적 환경', '복지 수준', '고용 안정성', '기업의 발전 가능성', '개인 성장 가능성']
    md_parts.append("## 13. 울산 지역 기업 인식 수준 (응답자 수)\n")
    header = "| 구 분 | 세부 항목 |"
    sep = "|:---|:---|"
    for cat in perception_cats:
        header += f" {cat} |"
        sep += " ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(271, 284):  # R272~R284
        row = data[i]
        if is_empty_row(row):
            break
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        line = f"| {a} | {b} |"
        for j in range(len(perception_cats)):
            val = row[2+j] if (2+j) < len(row) else None
            line += f" {fmt_val(val)} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 14. 울산 지역 기업 인식 수준 - 평균 (R286~R301)
    # ========================================
    md_parts.append("## 14. 울산 지역 기업 인식 수준 (평균)\n")
    header = "| 구 분 | 세부 항목 |"
    sep = "|:---|:---|"
    for cat in perception_cats:
        header += f" {cat} |"
        sep += " ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(288, 301):  # R289~R301
        row = data[i]
        if is_empty_row(row):
            break
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        line = f"| {a} | {b} |"
        for j in range(len(perception_cats)):
            val = row[2+j] if (2+j) < len(row) else None
            line += f" {fmt_val_raw(val)} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 15. 졸업 후 희망 취업 지역 (R303~R319)
    # ========================================
    cats15 = ['울산 지역 이유', '수도권(서울/경기/인천)', '경상권(대구, 경북, 부산, 경남)', '그 외 지역']
    md_parts.append("## 15. 졸업 후 희망 취업 지역\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats15:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(306, 319):  # R307~R319
        row = data[i]
        if is_empty_row(row):
            break
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats15)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 16. 울산 정주 환경이 타지역 취업에 미친 영향 (R321~R337)
    # ========================================
    cats16 = ['그렇다', '아니다']
    md_parts.append("## 16. 울산 정주 환경이 타지역 취업에 미친 영향\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats16:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(324, 337):  # R325~R337
        row = data[i]
        if is_empty_row(row):
            break
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats16)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 17. 타지역 취업에 영향을 미친 정주환경 요인(중복응답) (R339~R355)
    # ========================================
    cats17 = ['높은 주거비용 및 주택 부족', '대중교통 불편', '여가/문화생활 다양성 부족', '타지역 접근성 불편(KTX, 공항 등)', '공업도시 이미지로 인한 환경 우려', '청년 커뮤니티 및 네트워킹 기회 부족', '기타']
    md_parts.append("## 17. 타지역 취업에 영향을 미친 정주환경 요인(중복응답)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats17:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(342, 355):  # R343~R355
        row = data[i]
        if is_empty_row(row):
            break
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats17)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 18. 울산 지역 청년 정착 지원을 위한 정책(1순위) (R357~R373)
    # ========================================
    cats18 = ['청년 주택 공급/주거비 지원', '교통비 지원', '문화/여가 활동 지원', '청년 공간/커뮤니티 조성', '정착 지원금 제공', '생활 인프라 개선', '육아/보육 지원', '기타']
    md_parts.append("## 18. 울산 지역 청년 정착 지원을 위한 정책(1순위)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats18:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(360, 373):  # R361~R373
        row = data[i]
        if is_empty_row(row):
            break
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats18)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # ========================================
    # 19. 울산 지역 청년 정착 지원을 위한 정책(1순위+2순위) (R375~R391)
    # ========================================
    md_parts.append("## 19. 울산 지역 청년 정착 지원을 위한 정책(1순위+2순위)\n")
    header = "| 구 분 | 세부 항목 | 응답자 수 |"
    sep = "|:---|:---|---:|"
    for cat in cats18:
        header += f" {cat}(빈도) | {cat}(비율) |"
        sep += " ---:| ---:|"
    md_parts.append(header)
    md_parts.append(sep)
    for i in range(378, 391):  # R379~R391
        row = data[i]
        if is_empty_row(row):
            break
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c = fmt_val(row[2])
        line = f"| {a} | {b} | {c} |"
        for j in range(len(cats18)):
            fc = 3 + j*2
            rc = 4 + j*2
            freq = row[fc] if fc < len(row) and row[fc] else None
            rate = row[rc] if rc < len(row) and row[rc] else None
            f_str = fmt_val(freq) if freq else '-'
            r_str = f"{rate*100:.1f}%" if isinstance(rate, float) and 0 < abs(rate) <= 1 else ('-' if not rate else fmt_val(rate))
            line += f" {f_str} | {r_str} |"
        md_parts.append(line)
    md_parts.append("\n---\n")
    
    # 출력 파일 저장
    output_path = os.path.splitext(filepath)[0] + '.md'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(md_parts))
    
    print(f"Done: {output_path}")
    return output_path

if __name__ == '__main__':
    filepath = r'd:\git_rk\project\25_121_ulsan\output\(OUTPUT)울산 일자리 미스매치_260205_1822.xlsx'
    convert_to_markdown(filepath)
