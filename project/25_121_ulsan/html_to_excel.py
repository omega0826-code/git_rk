"""
HTML 파일의 표 데이터를 엑셀로 변환하는 스크립트
위치 기반(absolute positioning)으로 배치된 HTML 텍스트를 파싱하여 표 구조로 변환합니다.
"""

import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict
import os
import sys

def parse_html_to_table(html_file):
    """HTML 파일을 파싱하여 표 데이터 추출"""
    
    with open(html_file, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    soup = BeautifulSoup(html_content, 'html.parser')
    
    # 모든 div.pos 요소 찾기 (위치 정보가 있는 텍스트)
    divs = soup.find_all('div', class_='pos')
    
    # 위치별로 텍스트 저장 (top 좌표를 기준으로 행 구분)
    rows_data = defaultdict(list)
    
    for div in divs:
        style = div.get('style', '')
        
        # left, top 좌표 추출
        left_match = re.search(r'left:(\d+)px', style)
        top_match = re.search(r'top:(\d+)px', style)
        
        if left_match and top_match:
            left = int(left_match.group(1))
            top = int(top_match.group(1))
            
            # span 태그에서 텍스트 추출
            span = div.find('span')
            if span:
                text = span.get_text(strip=True)
                if text:  # 빈 텍스트 제외
                    rows_data[top].append((left, text))
    
    # top 좌표로 정렬하여 행 순서 결정
    sorted_tops = sorted(rows_data.keys())
    
    # 각 행의 데이터를 left 좌표로 정렬
    table_data = []
    for top in sorted_tops:
        row_items = sorted(rows_data[top], key=lambda x: x[0])
        row_text = [item[1] for item in row_items]
        table_data.append(row_text)
    
    return table_data

def merge_row_cells(table_data):
    """같은 행의 셀들을 병합하여 의미있는 컬럼으로 재구성"""
    
    # 표 구조 분석: 직종코드, 직종명, 전체 종사자 수, 2024년, 2025년, 2026년
    structured_data = []
    
    for row in table_data:
        if not row:
            continue
            
        # 행의 첫 번째 요소가 숫자(직종코드)인지 확인
        if row and row[0].replace(' ', '').isdigit():
            # 직종 데이터 행
            code = row[0]
            
            # 직종명 추출 (코드 다음부터 숫자나 '-'가 나오기 전까지)
            job_name_parts = []
            values = []
            
            for i in range(1, len(row)):
                item = row[i]
                # 숫자 또는 '-'이면 값으로 처리
                if item.replace(',', '').replace('-', '').isdigit() or item == '-':
                    values.append(item)
                else:
                    # 직종명의 일부
                    job_name_parts.append(item)
            
            job_name = ' '.join(job_name_parts)
            
            # 값이 4개 있어야 함 (전체, 2024, 2025, 2026)
            while len(values) < 4:
                values.append('-')
            
            structured_row = [code, job_name] + values[:4]
            structured_data.append(structured_row)
        else:
            # 헤더나 기타 행
            structured_data.append(row)
    
    return structured_data

def create_excel(table_data, output_file):
    """엑셀 파일 생성"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "울산지역 인력 및 훈련 수요조사"
    
    # 스타일 정의
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='맑은 고딕', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 작성
    headers = ['직종코드(KECO3)', '직종명', '전체 종사자 수', '2024년', '2025년', '2026년']
    ws.append(headers)
    
    # 헤더 스타일 적용
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 데이터 구조화
    structured_data = merge_row_cells(table_data)
    
    # 데이터 행 작성
    data_start_row = 2
    for row_data in structured_data:
        # 헤더 행이나 제목 행 건너뛰기
        if not row_data or len(row_data) < 2:
            continue
        
        # 첫 번째 요소가 숫자(직종코드)인 경우만 처리
        if row_data[0].replace(' ', '').isdigit():
            # 6개 컬럼에 맞춰 데이터 정리
            if len(row_data) >= 6:
                ws.append(row_data[:6])
            else:
                # 부족한 컬럼은 빈 값으로 채움
                padded_row = row_data + [''] * (6 - len(row_data))
                ws.append(padded_row[:6])
            
            # 스타일 적용
            current_row = ws.max_row
            for col_num in range(1, 7):
                cell = ws.cell(row=current_row, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                
                # 숫자 컬럼은 오른쪽 정렬
                if col_num >= 3:
                    cell.alignment = number_alignment
                else:
                    cell.alignment = data_alignment
    
    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 18  # 직종코드
    ws.column_dimensions['B'].width = 40  # 직종명
    ws.column_dimensions['C'].width = 15  # 전체 종사자 수
    ws.column_dimensions['D'].width = 12  # 2024년
    ws.column_dimensions['E'].width = 12  # 2025년
    ws.column_dimensions['F'].width = 12  # 2026년
    
    # 행 높이 조정
    ws.row_dimensions[1].height = 30
    
    # 엑셀 파일 저장
    wb.save(output_file)
    print(f"✅ 엑셀 파일이 생성되었습니다: {output_file}")

def main():
    # 입력 파일 경로
    if len(sys.argv) > 1:
        html_file = sys.argv[1]
    else:
        html_file = r'd:\git_rk\project\25_121_ulsan\data\075_115800.html'
    
    # 출력 파일 경로
    output_file = html_file.replace('.html', '.xlsx')
    
    print(f"📄 HTML 파일 읽기: {html_file}")
    
    # HTML 파싱
    table_data = parse_html_to_table(html_file)
    print(f"📊 {len(table_data)}개의 행을 추출했습니다.")
    
    # 엑셀 생성
    create_excel(table_data, output_file)
    
    print(f"\n✨ 변환 완료!")
    print(f"   입력: {html_file}")
    print(f"   출력: {output_file}")

if __name__ == '__main__':
    main()
